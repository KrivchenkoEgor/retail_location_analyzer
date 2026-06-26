"""
exporters/excel_report.py — генерация Excel-отчёта по результатам анализа.

Воспроизводит структуру bf_house_and_apartments_analysis.xlsx — 6 листов:
  1. Сводка        — основные метрики по каждой точке
  2. Демография    — residents по составу домохозяйств + итоги
  3. Детально      — конкуренты по категориям для каждой точки
  4. Выводы        — автоматические рекомендации
  5. Матрица сетей — присутствие ключевых сетей (✓/✗) по точкам
  6. Офисы/Образ.  — дополнительные источники трафика

build_report(results, output_path) — главная функция.
"""
from __future__ import annotations

from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config


# ---------------------------------------------------------------------------
# Стили
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="2F5597")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row=1):
    """Применить стиль к строке-заголовку."""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _autosize(ws, max_width=55):
    """Автоширина колонок по содержимому."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        width = min(max(lengths, default=10) + 2, max_width)
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Лист 1: Сводка
# ---------------------------------------------------------------------------
def _write_summary(wb, results):
    ws = wb.create_sheet("Сводка")
    headers = [
        "№", "Адрес", "Координаты", "Точность геокода",
        "Домов всего", "Жилых домов", "Квартир",
        "Конкурентов", "Офисов", "Уч. заведений",
        "Жителей (скорр.)", "Жит./конк.",
    ]
    ws.append(headers)
    _style_header(ws)

    for i, r in enumerate(results, 1):
        bld = r.get("buildings_summary", {})
        demo = r.get("demographics", {})
        if not r.get("ok"):
            ws.append([i, r.get("address", ""), "ОШИБКА", r.get("error", ""), "", "", "", "", "", "", "", ""])
            continue
        coord = f'{r["lat"]:.5f}, {r["lon"]:.5f}'
        ws.append([
            i, r["address"], coord, r.get("precision", ""),
            bld.get("total_buildings", 0), bld.get("residential", 0), bld.get("total_flats", 0),
            r.get("competitors_count", 0), len(r.get("poi", {}).get("offices", [])),
            len(r.get("poi", {}).get("education", [])),
            demo.get("corrected_residents", 0), r.get("residents_per_competitor", 0),
        ])

    # Итоговая строка
    last = ws.max_row + 1
    ws.cell(row=last, column=1, value="ИТОГО").font = BOLD
    totals = {
        5: sum(r.get("buildings_summary", {}).get("total_buildings", 0) for r in results if r.get("ok")),
        7: sum(r.get("buildings_summary", {}).get("total_flats", 0) for r in results if r.get("ok")),
        8: sum(r.get("competitors_count", 0) for r in results if r.get("ok")),
        9: sum(len(r.get("poi", {}).get("offices", [])) for r in results if r.get("ok")),
        10: sum(len(r.get("poi", {}).get("education", [])) for r in results if r.get("ok")),
        11: sum(r.get("demographics", {}).get("corrected_residents", 0) for r in results if r.get("ok")),
    }
    for col, val in totals.items():
        c = ws.cell(row=last, column=col, value=val)
        c.font = BOLD
    _autosize(ws)


# ---------------------------------------------------------------------------
# Лист 2: Демография
# ---------------------------------------------------------------------------
def _write_demographics(wb, results):
    ws = wb.create_sheet("Демография")
    headers = ["Адрес", "Квартир", "1 чел", "2 чел", "3 чел", "4 чел", "5+ чел",
               "Жителей (база)", "Жителей (скорр.)", "Среднее чел/кв"]
    ws.append(headers)
    _style_header(ws)

    for r in results:
        if not r.get("ok"):
            continue
        demo = r.get("demographics", {})
        hh = demo.get("households", {})
        ws.append([
            r["address"], demo.get("flats", 0),
            hh.get(1, 0), hh.get(2, 0), hh.get(3, 0), hh.get(4, 0), hh.get(5, 0),
            demo.get("base_residents", 0), demo.get("corrected_residents", 0),
            demo.get("final_avg_per_flat", 0),
        ])

    # Итого — прямое суммирование ключевых столбцов
    last = ws.max_row + 1
    ws.cell(row=last, column=1, value="ИТОГО").font = BOLD
    sums = {
        2: sum(r["demographics"]["flats"] for r in results if r.get("ok")),
        3: sum(r["demographics"]["households"].get(1, 0) for r in results if r.get("ok")),
        4: sum(r["demographics"]["households"].get(2, 0) for r in results if r.get("ok")),
        5: sum(r["demographics"]["households"].get(3, 0) for r in results if r.get("ok")),
        6: sum(r["demographics"]["households"].get(4, 0) for r in results if r.get("ok")),
        7: sum(r["demographics"]["households"].get(5, 0) for r in results if r.get("ok")),
        8: sum(r["demographics"]["base_residents"] for r in results if r.get("ok")),
        9: sum(r["demographics"]["corrected_residents"] for r in results if r.get("ok")),
    }
    for col, val in sums.items():
        ws.cell(row=last, column=col, value=val).font = BOLD
    _autosize(ws)


# ---------------------------------------------------------------------------
# Лист 3: Детально по точкам (конкуренты по категориям)
# ---------------------------------------------------------------------------
def _write_details(wb, results):
    ws = wb.create_sheet("Детально по точкам")
    headers = ["Адрес", "Категория", "Бренд/Название", "Тип (OSM)", "Расстояние, м"]
    ws.append(headers)
    _style_header(ws)

    for r in results:
        if not r.get("ok"):
            continue
        by_cat = r.get("shops_by_category", {})
        for cat, items in by_cat.items():
            for s in items:
                ws.append([
                    r["address"], cat, s["brand"], s["shop_type"],
                    s.get("distance_m", ""),
                ])
    _autosize(ws)


# ---------------------------------------------------------------------------
# Лист 4: Выводы и рекомендации
# ---------------------------------------------------------------------------
def _write_conclusions(wb, results):
    ws = wb.create_sheet("Выводы")
    ws.cell(row=1, column=1, value="Выводы и рекомендации").font = TITLE_FONT
    row = 3

    ok_results = [r for r in results if r.get("ok")]
    if not ok_results:
        ws.cell(row=row, column=1, value="Нет данных для анализа (все точки с ошибками).")
        return

    # Точка с минимальной конкуренцией (макс. жит./конк.)
    best = max(ok_results, key=lambda r: r.get("residents_per_competitor", 0))
    # Точка с максимальной конкуренцией
    worst = min(ok_results, key=lambda r: r.get("residents_per_competitor", float("inf")))
    # Точка с макс. жителей
    populous = max(ok_results, key=lambda r: r.get("demographics", {}).get("corrected_residents", 0))

    lines = [
        ("1. Лучший потенциал (мин. конкуренция на жителя):", ""),
        ("", f'{best["address"]} — {best.get("residents_per_competitor", 0)} жит./конк. '
             f'({best.get("competitors_count", 0)} конкурентов, '
             f'{best.get("demographics", {}).get("corrected_residents", 0)} жит.)'),
        ("", ""),
        ("2. Высокая конкуренция (перенасыщенный рынок):", ""),
        ("", f'{worst["address"]} — {worst.get("residents_per_competitor", 0)} жит./конк. '
             f'({worst.get("competitors_count", 0)} конкурентов)'),
        ("", ""),
        ("3. Самый большой жилой фонд:", ""),
        ("", f'{populous["address"]} — {populous.get("demographics", {}).get("corrected_residents", 0)} жит., '
             f'{populous.get("buildings_summary", {}).get("total_flats", 0)} кв.'),
        ("", ""),
        ("4. Топ-сети конкурентов по частоте присутствия:", ""),
    ]
    # Подсчёт присутствия сетей
    chain_counts = {}
    for r in ok_results:
        for brand in r.get("shop_brands", {}):
            chain_counts[brand] = chain_counts.get(brand, 0) + 1
    top_chains = sorted(chain_counts.items(), key=lambda kv: -kv[1])[:8]
    for brand, cnt in top_chains:
        lines.append(("", f"• {brand}: рядом с {cnt} из {len(ok_results)} точек"))

    for label, text in lines:
        if label:
            ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=2, value=text).alignment = WRAP
        row += 1
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 70


# ---------------------------------------------------------------------------
# Лист 5: Матрица конкуренции (сети × точки)
# ---------------------------------------------------------------------------
def _write_chain_matrix(wb, results):
    ws = wb.create_sheet("Матрица сетей")
    ok_results = [r for r in results if r.get("ok")]
    # Все бренды, отсортированные по частоте
    all_brands = {}
    for r in ok_results:
        for brand in r.get("shop_brands", {}):
            all_brands[brand] = all_brands.get(brand, 0) + 1
    brands = [b for b, _ in sorted(all_brands.items(), key=lambda kv: -kv[1])]

    headers = ["Сеть / Бренд"] + [f'Точка {i + 1}' for i in range(len(ok_results))] + ["Всего точек"]
    ws.append(headers)
    _style_header(ws)

    for brand in brands:
        row = [brand]
        total_points = 0
        for r in ok_results:
            cnt = r.get("shop_brands", {}).get(brand, 0)
            row.append("✓" if cnt > 0 else "—")
            if cnt > 0:
                total_points += 1
        row.append(total_points)
        ws.append(row)
    _autosize(ws)


# ---------------------------------------------------------------------------
# Лист 6: Офисы и учебные заведения
# ---------------------------------------------------------------------------
def _write_offices_education(wb, results):
    ws = wb.create_sheet("Офисы и уч. заведения")
    headers = ["Адрес", "Тип", "Название", "Расстояние, м"]
    ws.append(headers)
    _style_header(ws)

    for r in results:
        if not r.get("ok"):
            continue
        p = r.get("poi", {})
        for office in p.get("offices", []):
            ws.append([r["address"], "Офис", office["name"], office.get("distance_m", "")])
        for edu in p.get("education", []):
            ws.append([r["address"], "Образование", edu["name"], edu.get("distance_m", "")])
    _autosize(ws)


# ---------------------------------------------------------------------------
# Главная функция
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Лист 7: Данные МКД (my-gkh.ru)
# ---------------------------------------------------------------------------
def _write_mygkh(wb, results):
    ws = wb.create_sheet("Данные МКД")
    headers = [
        "Адрес", "Год постройки", "Этажей", "Жилых помещений",
        "Общая площадь, м²", "Серия", "Материал стен",
        "Управляющая компания", "Организация",
    ]
    ws.append(headers)
    _style_header(ws)

    for r in results:
        if not r.get("ok"):
            continue
        g = r.get("mygkh", {})
        if not g:
            continue
        ws.append([
            r["address"],
            g.get("year_built", ""),
            g.get("floors", ""),
            g.get("living_spaces", ""),
            g.get("total_area_sqm", ""),
            g.get("building_series", ""),
            g.get("building_material", ""),
            g.get("management_company", ""),
            g.get("organization_name", ""),
        ])

    # Характеристики и управление — отдельные строки под каждым адресом
    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="Детальные характеристики").font = TITLE_FONT
    row += 1
    for r in results:
        if not r.get("ok"):
            continue
        g = r.get("mygkh", {})
        if not g:
            continue
        chars = g.get("characteristics", {})
        if chars:
            ws.cell(row=row, column=1, value=r["address"]).font = BOLD
            row += 1
            for k, v in chars.items():
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=v)
                row += 1
        mgmt = g.get("management", {})
        if mgmt:
            ws.cell(row=row, column=1, value="Управление:").font = BOLD
            row += 1
            for k, v in mgmt.items():
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=v)
                row += 1
        utils = g.get("utility_providers", [])
        if utils:
            ws.cell(row=row, column=1, value="Поставщики ресурсов:").font = BOLD
            row += 1
            for u in utils:
                ws.cell(row=row, column=1, value=u)
                row += 1
        row += 1
    _autosize(ws)


def build_report(results: List[dict], output_path: str) -> str:
    """
    Сгенерировать Excel-отчёт по результатам analyze_many().

    results — список dict из analyzer.analyze_point/analyze_many.
    output_path — путь к .xlsx файлу.
    Возвращает output_path.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary(wb, results)
    _write_demographics(wb, results)
    _write_details(wb, results)
    _write_conclusions(wb, results)
    _write_chain_matrix(wb, results)
    _write_offices_education(wb, results)
    _write_mygkh(wb, results)

    wb.save(output_path)
    return output_path
